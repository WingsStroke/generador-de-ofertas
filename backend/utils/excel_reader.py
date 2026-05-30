import openpyxl
from openpyxl.utils import get_column_letter
from typing import List, Dict, Tuple, Optional
import re

# Nuevos componentes para manejo de esquemas variables
from utils.merged_cell_handler import MergedCellHandler
from utils.variable_header_detector import VariableHeaderDetector
from utils.inline_catalog_extractor import InlineCatalogExtractor, InlineCatalogEntry


class ExcelReader:
    """Lee archivos Excel y detecta la estructura del horario"""
    
    def __init__(self, file_path: str):
        self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        self.sheets = self.workbook.sheetnames
        self.current_sheet = self.workbook.active
        
    def set_sheet(self, sheet_name: str):
        """Cambia la hoja activa"""
        if sheet_name in self.sheets:
            self.current_sheet = self.workbook[sheet_name]
            return True
        return False
    
    def get_all_sheets(self) -> List[str]:
        """Obtiene nombres de todas las hojas"""
        return self.sheets
        
    def detect_all_schedule_structures(self) -> List[Tuple[int, int, List[str], List[Tuple[str, str, int]], int]]:
        """
        Detecta múltiples estructuras de horario en una hoja.
        Retorna lista de tuplas: (start_row, start_col, dias_encontrados, horas, end_row)
        """
        dias_semana = ["LUNES", "MARTES", "MIÉRCOLES", "MIERCOLES", "JUEVES", "VIERNES", "SÁBADO", "SABADO", "HORA"]

        def _matches_keyword(text: str, keyword: str) -> bool:
            return re.search(rf'\b{re.escape(keyword)}\b', text) is not None

        headers = []
        max_r = self.current_sheet.max_row
        
        for row_idx, row in enumerate(self.current_sheet.iter_rows(max_row=max_r), 1):
            row_count = 0
            first_match_col = None
            for col_idx, cell in enumerate(row, 1):
                if cell.value and isinstance(cell.value, str):
                    cell_upper = cell.value.strip().upper()
                    if any(_matches_keyword(cell_upper, dia) for dia in dias_semana):
                        row_count += 1
                        if first_match_col is None:
                            first_match_col = col_idx
            if row_count >= 2:
                headers.append((row_idx, first_match_col))
        
        if not headers:
            headers = [(1, 1)]
            
        structures = []
        for i, (start_row, start_col) in enumerate(headers):
            end_row = headers[i+1][0] - 1 if i + 1 < len(headers) else max_r
            
            dias_encontrados = []
            for cell in self.current_sheet[start_row]:
                if cell.value and isinstance(cell.value, str):
                    val_upper = cell.value.strip().upper()
                    for dia in ["LUNES", "MARTES", "MIÉRCOLES", "MIERCOLES", "JUEVES", "VIERNES", "SÁBADO", "SABADO"]:
                        if _matches_keyword(val_upper, dia):
                            dia_corto = self._dia_to_short(dia)
                            if dia_corto not in dias_encontrados:
                                dias_encontrados.append(dia_corto)
            
            if not dias_encontrados:
                dias_encontrados = ["L", "M", "W", "J", "V"]
                
            horas = self._extract_time_slots(start_row + 1, end_row)
            structures.append((start_row, start_col, dias_encontrados, horas, end_row))
            
        return structures

    def detect_schedule_structure(self) -> Tuple[int, int, List[str], List[Tuple[str, str, int]]]:
        """(Legacy) Devuelve una estructura combinada para compatibilidad"""
        structures = self.detect_all_schedule_structures()
        if not structures:
            return 1, 1, ["L", "M", "W", "J", "V"], []
            
        first_start_row, first_start_col, _, _, _ = structures[0]
        
        all_dias = []
        for s in structures:
            for d in s[2]:
                if d not in all_dias:
                    all_dias.append(d)
                    
        # Para evitar que tablas de la misma hoja mezclen las horas
        # (ej. 12:00 de tabla 2 quede después de 01:50 de tabla 1),
        # las convertimos a una escala de minutos monótona por tabla.
        horas_con_minutos = []
        for s in structures:
            last_mins = 0
            for h in s[3]:
                try:
                    hh, mm = map(int, h[0].split(':'))
                    mins = hh * 60 + mm
                    # Si el tiempo retrocede bruscamente (ej 12:00 -> 01:00), cruzamos al PM
                    while mins < last_mins - 240:
                        mins += 12 * 60
                    horas_con_minutos.append((h, mins))
                    last_mins = mins
                except Exception:
                    horas_con_minutos.append((h, 0))
                    
        # Ordenar cronológicamente
        horas_con_minutos.sort(key=lambda x: x[1])
        
        all_horas = []
        seen_horas = set()
        for h, _ in horas_con_minutos:
            k = (h[0], h[1])
            if k not in seen_horas:
                seen_horas.add(k)
                all_horas.append(h)
                    
        return first_start_row, first_start_col, all_dias, all_horas
    
    def _dia_to_short(self, dia: str) -> str:
        """Convierte nombre de día completo a abreviatura"""
        mapping = {
            "LUNES": "L",
            "MARTES": "M",
            "MIÉRCOLES": "W",
            "MIERCOLES": "W",
            "JUEVES": "J",
            "VIERNES": "V",
            "SÁBADO": "S",
            "SABADO": "S"
        }
        return mapping.get(dia, dia[0])
    
    def _extract_time_slots(self, start_row: int, end_row: int = None) -> List[Tuple[str, str, int]]:
        """Extrae las franjas horarias detectadas con la fila Excel real donde viven.

        Maneja:
          - Etiqueta normal "7:00 - 7:50" en una sola celda
          - Etiqueta partida en 2 celdas consecutivas: "12:00 -" + "12:50"
          - Etiquetas multilínea "9:30 -\n10:20"
          - Filas de footer/descripción son ignoradas (texto largo, sin patrón de hora)
        """
        time_pattern = re.compile(r'(\d{1,2})[:\s]*(\d{2})')
        max_row = self.current_sheet.max_row
        if end_row is not None:
            max_row = min(max_row, end_row)

        rows_data = []
        for r in range(start_row, max_row + 1):
            cell_value = self.current_sheet.cell(row=r, column=1).value
            if isinstance(cell_value, str):
                stripped = cell_value.strip()
                matches = time_pattern.findall(stripped)
                # Filtrar filas tipo footer: texto largo y muchas palabras
                is_footer = len(stripped) > 25 and len(stripped.split()) > 4
                rows_data.append((r, stripped, matches, is_footer))
            else:
                rows_data.append((r, None, [], False))

        horas: List[Tuple[str, str, int]] = []
        i = 0
        while i < len(rows_data):
            r, val, matches, is_footer = rows_data[i]
            if is_footer:
                i += 1
                continue
            if len(matches) >= 2:
                inicio = f"{matches[0][0].zfill(2)}:{matches[0][1]}"
                fin = f"{matches[1][0].zfill(2)}:{matches[1][1]}"
                horas.append((inicio, fin, r))
                i += 1
            elif len(matches) == 1 and i + 1 < len(rows_data):
                # Caso "12:00 -" + "12:50" en filas consecutivas
                r2, val2, matches2, is_footer2 = rows_data[i + 1]
                if not is_footer2 and len(matches2) >= 1:
                    inicio = f"{matches[0][0].zfill(2)}:{matches[0][1]}"
                    fin = f"{matches2[0][0].zfill(2)}:{matches2[0][1]}"
                    horas.append((inicio, fin, r))
                    i += 2
                else:
                    i += 1
            else:
                i += 1

        # De-duplicar por hora_inicio (mantener primera aparición)
        seen = set()
        unique = []
        for h in horas:
            key = (h[0], h[1])
            if key not in seen:
                seen.add(key)
                unique.append(h)

        if not unique:
            unique = [
                ("07:00", "07:50", start_row), ("07:50", "08:40", start_row + 1),
                ("08:40", "09:30", start_row + 2), ("09:30", "10:20", start_row + 3),
            ]

        return unique
    
    def get_cell_content(self, row: int, col: int) -> Optional[str]:
        """Obtiene el contenido de una celda"""
        cell = self.current_sheet.cell(row=row, column=col)
        return str(cell.value).strip() if cell.value else None
    
    def get_merged_cells(self) -> List[Dict]:
        """Obtiene información de celdas fusionadas"""
        merged_cells = []
        for merged_range in self.current_sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            merged_cells.append({
                "ref": str(merged_range),
                "min_row": min_row,
                "min_col": min_col,
                "max_row": max_row,
                "max_col": max_col,
                "rowspan": max_row - min_row + 1,
                "colspan": max_col - min_col + 1
            })
        return merged_cells
    
    def extract_schedule_cells(self, use_merged_handler: bool = True) -> List[Dict]:
        """Extrae todas las celdas relevantes del horario iterando sobre todas las sub-tablas."""
        structures = self.detect_all_schedule_structures()
        cells_data = []
        dias_keywords = ["LUNES", "MARTES", "MIÉRCOLES", "MIERCOLES", "JUEVES", "VIERNES", "SÁBADO", "SABADO"]

        merged_handler = None
        if use_merged_handler:
            merged_handler = MergedCellHandler(self.current_sheet)

        for start_row, start_col, dias, horas, end_row in structures:
            if not horas:
                continue

            days_end_col = start_col + len(dias)
            hora_by_row = {h[2]: (h[0], h[1]) for h in horas}
            first_time_row = horas[0][2]
            last_time_row = horas[-1][2]
            scan_end = min(last_time_row + 2, end_row)

            current_hora = None
            for r in range(first_time_row, scan_end + 1):
                if r in hora_by_row:
                    current_hora = hora_by_row[r]

                if current_hora is None:
                    continue

                if merged_handler:
                    a_val = merged_handler.get_effective_value(r, start_col)
                else:
                    a_val = self.current_sheet.cell(row=r, column=start_col).value
                    
                if r not in hora_by_row and isinstance(a_val, str):
                    stripped = a_val.strip()
                    upper_strip = stripped.upper()
                    if upper_strip.startswith('SEMESTRE') or upper_strip.startswith('UNIVERSIDAD') or 'PROGRAMA' in upper_strip or upper_strip == 'CURSOS LIBRES':
                        break
                    if len(stripped) > 25 and len(stripped.split()) > 4:
                        break

                for dia_idx, dia in enumerate(dias):
                    col = start_col + 1 + dia_idx
                    if col > days_end_col:
                        break
                    
                    if merged_handler:
                        content = merged_handler.get_effective_value(r, col)
                        if content:
                            content = str(content).strip()
                    else:
                        content = self.get_cell_content(r, col)

                    if content and content.lower() not in ['none', 'nan', '']:
                        is_day_header = any(d in content.upper() for d in dias_keywords)
                        if is_day_header:
                            continue

                        cell_ref = f"{get_column_letter(col)}{r}"
                        cells_data.append({
                            "dia": dia,
                            "hora_inicio": current_hora[0],
                            "hora_fin": current_hora[1],
                            "texto": content,
                            "celda_ref": cell_ref,
                            "row": r,
                            "col": col
                        })

        return cells_data

    def detect_catalog(self):
        """Wrapper sobre catalog_reader.detect_catalog usando la columna fin del horario."""
        from utils.catalog_reader import detect_catalog as _dc
        _, start_col, dias, _ = self.detect_schedule_structure()
        days_end_col = start_col + len(dias)
        return _dc(self.current_sheet, days_end_col)

    def read_catalog_entries(self, catalog: Dict) -> List[Dict]:
        from utils.catalog_reader import read_catalog_entries as _rce
        return _rce(self.current_sheet, catalog)
    
    def get_preview_grid(self, max_rows: int = 50, max_cols: int = 10) -> List[Dict]:
        """Genera una representación del Excel para preview"""
        preview_cells = []
        merged_info = {str(m["ref"]): m for m in self.get_merged_cells()}
        
        for row_idx in range(1, min(max_rows + 1, self.current_sheet.max_row + 1)):
            for col_idx in range(1, min(max_cols + 1, self.current_sheet.max_column + 1)):
                cell = self.current_sheet.cell(row=row_idx, column=col_idx)
                cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
                
                is_merged = False
                rowspan = 1
                colspan = 1
                
                for merged_range, info in merged_info.items():
                    if (info["min_row"] <= row_idx <= info["max_row"] and 
                        info["min_col"] <= col_idx <= info["max_col"]):
                        is_merged = True
                        if row_idx == info["min_row"] and col_idx == info["min_col"]:
                            rowspan = info["rowspan"]
                            colspan = info["colspan"]
                        break
                
                preview_cells.append({
                    "ref": cell_ref,
                    "value": str(cell.value) if cell.value else None,
                    "row": row_idx,
                    "col": col_idx,
                    "rowspan": rowspan,
                    "colspan": colspan,
                    "is_merged": is_merged
                })
        
        return preview_cells
    
    def detect_header_row_adaptive(self) -> int:
        """
        Detecta la fila de headers usando VariableHeaderDetector.
        
        Returns:
            Número de fila con los headers (1-based)
        """
        detector = VariableHeaderDetector()
        return detector.detect_header_row(self.current_sheet)
    
    def detect_inline_catalog(self, header_row: int = None) -> Optional[Dict]:
        """
        Detecta si existe un catálogo inline en la hoja.
        
        Args:
            header_row: Fila de headers (si None, se detecta automáticamente)
            
        Returns:
            Dict con estructura del catálogo o None
        """
        if header_row is None:
            header_row = self.detect_header_row_adaptive()
        
        extractor = InlineCatalogExtractor()
        return extractor.detect_catalog_structure(self.current_sheet, header_row)
    
    def extract_inline_catalog(self, header_row: int = None) -> List[InlineCatalogEntry]:
        """
        Extrae el catálogo inline iterando sobre todas las sub-tablas de la hoja.
        
        Args:
            header_row: (Legacy) Ignorado, se autodetectan los headers por sub-tabla.
            
        Returns:
            Lista de entradas del catálogo combinadas
        """
        structures = self.detect_all_schedule_structures()
        merged_handler = MergedCellHandler(self.current_sheet)
        extractor = InlineCatalogExtractor(merged_handler)
        
        all_entries = []
        for start_row, start_col, dias, horas, end_row in structures:
            structure = extractor.detect_catalog_structure(self.current_sheet, start_row)
            if structure:
                entries = extractor.extract_catalog(
                    self.current_sheet, 
                    start_row, 
                    structure,
                    max_empty_rows=10, # Aumentado para tolerar huecos entre tablas
                    end_row=end_row
                )
                all_entries.extend(entries)
        
        return extractor.deduplicate_entries(all_entries)
    
    def get_merged_cell_handler(self) -> MergedCellHandler:
        """Retorna un MergedCellHandler para la hoja actual."""
        return MergedCellHandler(self.current_sheet)
    
    def close(self):
        """Cierra el workbook"""
        if self.workbook:
            self.workbook.close()
