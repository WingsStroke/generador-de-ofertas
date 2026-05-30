"""
excel_html_renderer.py
Convierte una hoja de openpyxl a HTML estilizado preservando:
  - Tamaños de fuente, negrita, cursiva, color de texto
  - Color de fondo de celda
  - Celdas fusionadas (rowspan/colspan)
  - Anchos de columna proporcionales
  - Alineación de texto
  - Resaltado de la celda activa (parámetro highlight_refs)
"""
from __future__ import annotations

import html
from typing import Optional
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles.colors import COLOR_INDEX, aRGB_REGEX


# ── Helpers de color ────────────────────────────────────────────────────────

_INDEXED_COLORS = [
    "000000", "FFFFFF", "FF0000", "00FF00", "0000FF", "FFFF00", "FF00FF",
    "00FFFF", "000000", "FFFFFF", "FF0000", "00FF00", "0000FF", "FFFF00",
    "FF00FF", "00FFFF", "800000", "008000", "000080", "808000", "800080",
    "008080", "C0C0C0", "808080", "9999FF", "993366", "FFFFCC", "CCFFFF",
    "660066", "FF8080", "0066CC", "CCCCFF", "000080", "FF00FF", "FFFF00",
    "00FFFF", "800080", "800000", "008080", "0000FF", "00CCFF", "CCFFFF",
    "CCFFCC", "FFFF99", "99CCFF", "FF99CC", "CC99FF", "FFCC99", "3366FF",
    "33CCCC", "99CC00", "FFCC00", "FF9900", "FF6600", "666699", "969696",
    "003366", "339966", "003300", "333300", "993300", "993366", "333399",
    "333333",
]


def _resolve_color(color) -> Optional[str]:
    """Devuelve un código hex #RRGGBB o None si el color es transparente/auto."""
    if color is None:
        return None
    try:
        t = color.type
    except Exception:
        return None

    if t == "rgb":
        rgb = color.rgb  # 8 chars AARRGGBB
        if rgb in ("00000000", "FFFFFFFF", "00FFFFFF"):
            return None
        return f"#{rgb[2:]}"  # Skip alpha
    if t == "indexed":
        idx = color.indexed
        if idx is None or idx >= len(_INDEXED_COLORS) or idx in (64, 65):
            return None
        return f"#{_INDEXED_COLORS[idx]}"
    if t == "theme":
        # Theme colors are complex — skip for safety
        return None
    return None


# ── Ancho de columna ────────────────────────────────────────────────────────

def _col_width_px(sheet, col_idx: int) -> int:
    """Convierte el ancho de columna de openpyxl (caracteres) a píxeles aprox."""
    col_letter = get_column_letter(col_idx)
    dim = sheet.column_dimensions.get(col_letter)
    if dim and dim.width:
        # ~7px por unidad de ancho de Excel
        return max(30, int(dim.width * 7))
    return 80  # default


def _row_height_px(sheet, row_idx: int) -> int:
    """Convierte la altura de fila de openpyxl (puntos) a píxeles aprox."""
    dim = sheet.row_dimensions.get(row_idx)
    if dim and dim.height:
        return max(16, int(dim.height * 1.33))  # pt → px aprox
    return 20  # default


# ── Renderizador principal ──────────────────────────────────────────────────

def sheet_to_html(
    workbook: openpyxl.Workbook,
    sheet_name: str,
    highlight_ref: Optional[str] = None,
    max_rows: Optional[int] = None,
    max_cols: Optional[int] = None,
) -> str:
    """
    Convierte una hoja de openpyxl en HTML estilizado.

    Args:
        workbook: Workbook ya abierto (data_only=True recomendado).
        sheet_name: Nombre de la hoja a renderizar.
        highlight_ref: Referencia de celda a resaltar (ej. "B5"). Opcional.
        max_rows: Limitar filas renderizadas (None = todas).
        max_cols: Limitar columnas renderizadas (None = todas).

    Returns:
        String HTML completo con estilos inline.
    """
    if sheet_name not in workbook.sheetnames:
        return f"<p>Hoja '{html.escape(sheet_name)}' no encontrada.</p>"

    ws = workbook[sheet_name]

    n_rows = ws.max_row or 1
    n_cols = ws.max_column or 1
    if max_rows:
        n_rows = min(n_rows, max_rows)
    if max_cols:
        n_cols = min(n_cols, max_cols)

    # Construir mapa de celdas fusionadas: (row, col) → (min_row, min_col, rowspan, colspan)
    merge_map: dict[tuple, tuple | None] = {}
    for merged_range in ws.merged_cells.ranges:
        r1, c1, r2, c2 = (
            merged_range.min_row, merged_range.min_col,
            merged_range.max_row, merged_range.max_col,
        )
        rowspan = r2 - r1 + 1
        colspan = c2 - c1 + 1
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                if r == r1 and c == c1:
                    merge_map[(r, c)] = (r1, c1, rowspan, colspan)
                else:
                    merge_map[(r, c)] = None  # celda absorbida

    parts = [
        '<div class="xlsx-html-preview">',
        '<table class="xlsx-html-table" cellspacing="0" cellpadding="0">',
        '<tbody>',
    ]

    for r in range(1, n_rows + 1):
        row_h = _row_height_px(ws, r)
        parts.append(f'<tr style="height:{row_h}px">')

        for c in range(1, n_cols + 1):
            merge_info = merge_map.get((r, c), "NORMAL")

            # Celda absorbida → saltar
            if merge_info is None:
                continue

            # Determinar rowspan/colspan
            if merge_info == "NORMAL":
                rowspan, colspan = 1, 1
            else:
                _, _, rowspan, colspan = merge_info

            # Limitar spans al área renderizada
            rowspan = min(rowspan, n_rows - r + 1)
            colspan = min(colspan, n_cols - c + 1)

            cell = ws.cell(row=r, column=c)
            cell_ref = f"{get_column_letter(c)}{r}"

            # ── Extraer estilos ────────────────────────────────────────────
            style_parts = []

            # Fondo
            try:
                fill = cell.fill
                if fill and fill.patternType not in (None, "none"):
                    bg = _resolve_color(fill.fgColor)
                    if bg:
                        style_parts.append(f"background-color:{bg}")
            except Exception:
                pass

            # Fuente
            font_size = 11
            try:
                font = cell.font
                if font:
                    if font.size:
                        # Umbral mínimo de legibilidad: 11pt
                        font_size = max(11, int(font.size))
                    if font.bold:
                        style_parts.append("font-weight:bold")
                    if font.italic:
                        style_parts.append("font-style:italic")
                    fc = _resolve_color(font.color)
                    if fc:
                        style_parts.append(f"color:{fc}")
            except Exception:
                pass

            style_parts.append(f"font-size:{font_size}pt")

            # Alineación
            try:
                align = cell.alignment
                if align:
                    h_align = align.horizontal
                    if h_align in ("center", "right", "left"):
                        style_parts.append(f"text-align:{h_align}")
                    v_align = align.vertical
                    if v_align == "center":
                        style_parts.append("vertical-align:middle")
                    elif v_align == "bottom":
                        style_parts.append("vertical-align:bottom")
                    if align.wrap_text:
                        style_parts.append("white-space:pre-wrap")
            except Exception:
                pass

            # Borde (solo detectar si existe)
            try:
                border = cell.border
                border_css = []
                sides = [("top", border.top), ("right", border.right),
                         ("bottom", border.bottom), ("left", border.left)]
                for side_name, side in sides:
                    if side and side.style:
                        border_css.append(f"border-{side_name}:1px solid #999")
                style_parts.extend(border_css)
            except Exception:
                pass

            # ── Resaltado de celda activa ──────────────────────────────────
            is_highlighted = highlight_ref and cell_ref == highlight_ref

            # ── Valor de la celda ──────────────────────────────────────────
            raw_value = cell.value
            if raw_value is None:
                display = ""
            else:
                display = html.escape(str(raw_value).strip())

            # ── Armar atributos del <td> ───────────────────────────────────
            style_str = ";".join(style_parts)
            attrs = [f'data-ref="{cell_ref}"']
            if rowspan > 1:
                attrs.append(f'rowspan="{rowspan}"')
            if colspan > 1:
                attrs.append(f'colspan="{colspan}"')
            if style_str:
                attrs.append(f'style="{style_str}"')
            if is_highlighted:
                attrs.append('class="xlsx-cell-highlight"')

            parts.append(f'<td {" ".join(attrs)}>{display}</td>')

        parts.append("</tr>")

    parts.extend(["</tbody>", "</table>", "</div>"])
    return "\n".join(parts)
