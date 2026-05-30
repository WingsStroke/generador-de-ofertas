"""
InlineCatalogExtractor - Extrae catálogo cuando está inline en la misma hoja
"""
from typing import List, Dict, Optional, Any
from openpyxl.utils import get_column_letter


class InlineCatalogEntry:
    """Representa una entrada del catálogo inline."""
    
    def __init__(self, nombre: str, horas: Any = None, codigo: str = None, 
                 grupo: str = None, fila_excel: int = None, docente: str = None):
        self.nombre = nombre
        self.horas = horas
        self.codigo = codigo
        self.grupo = grupo
        self.fila_excel = fila_excel
        self.docente = docente
    
    def to_dict(self) -> dict:
        return {
            'nombre': self.nombre,
            'horas': self.horas,
            'codigo': self.codigo,
            'grupo': self.grupo,
            'fila_excel': self.fila_excel,
            'docente': self.docente
        }
    
    def __repr__(self):
        return f"InlineCatalogEntry({self.nombre[:30]}, codigo={self.codigo}, grupo={self.grupo})"


class InlineCatalogExtractor:
    """
    Extrae catálogo de asignaturas cuando está en la misma hoja que el horario.
    Común en archivos de Ingeniería de Alimentos (columnas 8-11).
    """
    
    # Patrones conocidos de catálogo inline
    CATALOG_PATTERNS = [
        {
            'name': 'alimentos_standard',
            'nombre_col': 8,
            'horas_col': 9,
            'codigo_col': 10,
            'grupo_col': 11,
            'docente_col': 12,
        },
        {
            'name': 'minimal_catalog',
            'nombre_col': 8,
            'horas_col': 9,
            'codigo_col': None,
            'grupo_col': None,
            'docente_col': None,
        }
    ]
    
    def __init__(self, merged_cell_handler=None):
        self.merged_handler = merged_cell_handler
    
    def detect_catalog_structure(self, sheet, header_row: int) -> Optional[dict]:
        """
        Detecta si existe un catálogo inline y retorna su estructura.
        
        Args:
            sheet: Hoja de Excel
            header_row: Fila donde están los headers
            
        Returns:
            Dict con estructura o None si no hay catálogo
        """
        structure = {
            'nombre_col': None,
            'horas_col': None,
            'codigo_col': None,
            'grupo_col': None,
            'docente_col': None,
            'pattern_name': None
        }
        
        # Buscar en columnas 7-12 (donde típicamente está el catálogo)
        for col_idx in range(7, min(13, sheet.max_column + 1)):
            cell = sheet.cell(row=header_row, column=col_idx)
            if not cell.value:
                continue
            
            val_upper = str(cell.value).upper().strip()
            
            # 1. Detectar columna de docente PRIMERO (evita conflicto con 'NOMBRE')
            if any(keyword in val_upper for keyword in ['DOCENTE', 'PROFESOR']):
                structure['docente_col'] = col_idx
            
            # 2. Detectar columna de codigo PRIMERO (evita conflicto con 'ASIGNATURA')
            elif any(keyword in val_upper for keyword in ['CODIGO', 'CÓDIGO']):
                structure['codigo_col'] = col_idx
            
            # 3. Detectar columna de horas
            elif val_upper in ['HORAS', 'HORAS SEMANALES'] or 'HORA' in val_upper:
                structure['horas_col'] = col_idx
            
            # 4. Detectar columna de grupo
            elif val_upper == 'GRUPO' or val_upper == 'GRUPOS':
                structure['grupo_col'] = col_idx
            
            # 5. Detectar columna de nombre
            elif any(keyword in val_upper for keyword in ['ASIGNATURA', 'MATERIA', 'NOMBRE']):
                structure['nombre_col'] = col_idx
                structure['pattern_name'] = 'detected'
        
        # Validar: debe tener al menos columna de nombre
        if structure['nombre_col']:
            return structure
        
        return None
    
    def extract_catalog(self, sheet, header_row: int, 
                        structure: dict = None,
                        max_empty_rows: int = 3,
                        end_row: int = None) -> List[InlineCatalogEntry]:
        """
        Extrae entradas del catálogo inline.
        
        Args:
            sheet: Hoja de Excel
            header_row: Fila de headers
            structure: Estructura del catálogo (si None, se detecta automáticamente)
            max_empty_rows: Máximo de filas vacías antes de detener
            end_row: Fila máxima a leer (útil para multi-tabla)
            
        Returns:
            Lista de InlineCatalogEntry
        """
        if structure is None:
            structure = self.detect_catalog_structure(sheet, header_row)
        
        if not structure:
            return []
        
        entries = []
        empty_count = 0
        row = header_row + 1
        
        nombre_col = structure.get('nombre_col')
        horas_col = structure.get('horas_col')
        codigo_col = structure.get('codigo_col')
        grupo_col = structure.get('grupo_col')
        docente_col = structure.get('docente_col')
        
        max_r = end_row if end_row is not None else sheet.max_row
        while row <= max_r and empty_count < max_empty_rows:
            # Obtener nombre usando merged_handler si está disponible
            if self.merged_handler:
                nombre = self.merged_handler.get_effective_value(row, nombre_col)
            else:
                nombre = sheet.cell(row=row, column=nombre_col).value
            
            # Limpiar y validar nombre
            if nombre:
                nombre_str = str(nombre).strip()
                
                # Reemplazar saltos de línea y múltiples espacios por un solo espacio
                nombre_str = " ".join(nombre_str.split())
                
                nombre_upper = nombre_str.upper()
                
                # Ignorar filas que se ven como headers o están vacías
                es_header = any(h in nombre_upper for h in [
                    'ASIGNATURA', 'SEMESTRE', 'ELECTIVA', 'CURSO LIBRE', 'CURSOS LIBRES'
                ])
                
                if nombre_str and len(nombre_str) > 2 and not es_header:
                    
                    # Extraer otras columnas
                    horas = None
                    codigo = None
                    grupo = None
                    
                    if horas_col:
                        horas_cell = sheet.cell(row=row, column=horas_col).value
                        if horas_cell:
                            try:
                                horas = int(horas_cell)
                            except (ValueError, TypeError):
                                horas = str(horas_cell).strip()
                    
                    if codigo_col:
                        codigo_cell = sheet.cell(row=row, column=codigo_col).value
                        if codigo_cell:
                            codigo = str(codigo_cell).strip()
                    
                    if grupo_col:
                        grupo_cell = sheet.cell(row=row, column=grupo_col).value
                        if grupo_cell:
                            grupo = str(grupo_cell).strip()
                    
                    docente = None
                    if docente_col:
                        docente_cell = sheet.cell(row=row, column=docente_col).value
                        if docente_cell:
                            docente = str(docente_cell).strip()
                    
                    entry = InlineCatalogEntry(
                        nombre=nombre_str,
                        horas=horas,
                        codigo=codigo,
                        grupo=grupo,
                        fila_excel=row,
                        docente=docente
                    )
                    entries.append(entry)
                    empty_count = 0
                else:
                    empty_count += 1
            else:
                empty_count += 1
            
            row += 1
        
        return entries
    
    def create_lookup_index(self, entries: List[InlineCatalogEntry]) -> Dict[str, InlineCatalogEntry]:
        """
        Crea un índice de búsqueda por nombre de materia.
        
        Útil para enriquecer bloques de horario con datos del catálogo.
        """
        index = {}
        
        for entry in entries:
            # Índice por nombre exacto
            index[entry.nombre.upper()] = entry
            
            # Índice por primera palabra (para búsqueda parcial)
            first_word = entry.nombre.split()[0].upper()
            if first_word not in index:
                index[first_word] = entry
        
        return index
    
    def find_match(self, text: str, index: Dict[str, InlineCatalogEntry], 
                   threshold_chars: int = 5) -> Optional[InlineCatalogEntry]:
        """
        Busca una coincidencia aproximada en el catálogo.
        
        Args:
            text: Texto a buscar (ej: "Cálculo Diferencial A1")
            index: Índice creado con create_lookup_index
            threshold_chars: Mínimo de caracteres para considerar match
            
        Returns:
            InlineCatalogEntry coincidente o None
        """
        text_upper = text.upper().strip()
        
        # Búsqueda exacta
        if text_upper in index:
            return index[text_upper]
        
        # Búsqueda por contención
        for name, entry in index.items():
            if len(name) >= threshold_chars:
                if name in text_upper or text_upper in name:
                    return entry
        
        # Búsqueda por primera palabra
        first_word = text_upper.split()[0] if text_upper else ''
        if first_word and len(first_word) >= 3:
            if first_word in index:
                return index[first_word]
        
        return None
    
    def deduplicate_entries(self, entries: List[InlineCatalogEntry]) -> List[InlineCatalogEntry]:
        """
        Elimina duplicados del catálogo, pero respeta grupos diferentes para la misma materia.
        """
        seen = set()
        unique = []
        
        for entry in entries:
            grp = entry.grupo.upper() if entry.grupo else ""
            key = (entry.nombre.upper(), grp)
            if key not in seen:
                seen.add(key)
                unique.append(entry)
        
        return unique
    
    def __repr__(self):
        return "InlineCatalogExtractor()"
