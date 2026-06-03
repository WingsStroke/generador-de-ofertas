"""
pdf_converter.py
================
Convierte un archivo PDF de oferta académica a un archivo XLSX temporal
que puede ser procesado por el pipeline existente (ScheduleProcessor / ExcelReader).

Algoritmo:
  1. Para cada página del PDF, pdfplumber detecta las tablas.
  2. Se extraen todas las celdas de TODAS las tablas de esa página (soporte multi-tabla).
  3. Las coordenadas X e Y se agrupan con tolerancia para crear una cuadrícula uniforme.
  4. Cada celda se ubica en la cuadrícula y su texto se escribe en la hoja Excel.
  5. Se aplican las fusiones (merged cells) según el span de la celda en la cuadrícula.

Artefactos de texto conocidos (solapamientos físicos en el PDF) se corrigen con
`_clean_pdf_text` antes de escribir al XLSX.
"""

import re
import tempfile
import os
import logging
from pathlib import Path
from typing import List, Optional

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# Correcciones de texto conocidas (artefactos de solapamiento en el PDF fuente)
# ──────────────────────────────────────────────────────────────────────────────
_PDF_TEXT_CORRECTIONS = {
    "ASccailoundes": "Acciones",
    "AIngtrrooidnudcrcuisón": "Introducción",
}


def _clean_pdf_text(text: str) -> str:
    """Aplica correcciones de texto conocidas a artefactos del PDF."""
    if not text:
        return ""
    for bad, good in _PDF_TEXT_CORRECTIONS.items():
        text = text.replace(bad, good)
    return text


# ──────────────────────────────────────────────────────────────────────────────
# Agrupamiento de coordenadas
# ──────────────────────────────────────────────────────────────────────────────
def _cluster_coordinates(coords: List[float], tolerance: float = 3.0) -> List[float]:
    """
    Agrupa coordenadas cercanas en un único valor representativo (el promedio del grupo).
    Esto normaliza pequeñas diferencias de coordenadas entre celdas adyacentes.
    """
    if not coords:
        return []
    coords = sorted(set(coords))
    clusters: List[List[float]] = []
    current: List[float] = [coords[0]]

    for val in coords[1:]:
        if val - current[0] <= tolerance:
            current.append(val)
        else:
            clusters.append(current)
            current = [val]
    clusters.append(current)

    return [sum(c) / len(c) for c in clusters]


def _find_grid_index(val: float, coord_list: List[float], tolerance: float = 3.0) -> int:
    """Encuentra el índice más cercano en una lista de coordenadas agrupadas."""
    for idx, coord in enumerate(coord_list):
        if abs(coord - val) <= tolerance:
            return idx
    # Fallback: índice del más cercano
    return min(range(len(coord_list)), key=lambda i: abs(coord_list[i] - val))


# ──────────────────────────────────────────────────────────────────────────────
# Conversión principal
# ──────────────────────────────────────────────────────────────────────────────
def pdf_color_to_hex(color) -> Optional[str]:
    """
    Convierte un color de pdfplumber (float o lista de floats) a un string hexadecimal RRGGBB.
    """
    if not color:
        return None
    if isinstance(color, (int, float)):
        val = int(color * 255) if color <= 1.0 else int(color)
        val = max(0, min(255, val))
        return f"{val:02X}{val:02X}{val:02X}"
    if isinstance(color, (list, tuple)):
        rgb = []
        for c in color[:3]:
            val = int(c * 255) if c <= 1.0 else int(c)
            val = max(0, min(255, val))
            rgb.append(val)
        while len(rgb) < 3:
            rgb.append(rgb[-1] if rgb else 0)
        return f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"
    return None


def pdf_to_xlsx(pdf_path: str) -> str:
    """
    Convierte un PDF de oferta académica a un archivo XLSX temporal.

    Parámetros
    ----------
    pdf_path : str
        Ruta al archivo PDF de entrada.

    Retorna
    -------
    str
        Ruta al archivo XLSX generado en un directorio temporal del sistema.
        El llamador es responsable de eliminarlo cuando ya no sea necesario.

    Lanza
    -----
    ImportError
        Si `pdfplumber` u `openpyxl` no están instalados.
    ValueError
        Si el PDF no contiene tablas en ninguna página.
    RuntimeError
        Si ocurre cualquier otro error durante la conversión.
    """
    try:
        import pdfplumber  # noqa: F401
    except ImportError:
        raise ImportError(
            "La librería 'pdfplumber' no está instalada. "
            "Ejecuta: pip install pdfplumber"
        )

    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "La librería 'openpyxl' no está instalada. "
            "Ejecuta: pip install openpyxl"
        )

    import pdfplumber

    # Tolerancias de clustering (validadas contra los PDFs de la universidad)
    # NOTA: no aumentar por encima de 5/8 — con valores mayores las filas
    # adyacentes se colapsan y se pierden materias en páginas con multi-tabla.
    X_TOLERANCE = 3.0   # Para columnas
    Y_TOLERANCE = 5.0   # Para filas
    CROP_PADDING = 1.5  # Puntos de recorte interior para extracción de texto

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Quitar hoja por defecto

    sheets_created = 0

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_idx, page in enumerate(pdf.pages):
                page_num = page_idx + 1

                # Configuración de extracción de tablas mejorada
                table_settings_lines = {
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines",
                    "snap_x_tolerance": 3,
                    "snap_y_tolerance": 3,
                    "join_tolerance": 4,
                }
                tables = page.find_tables(table_settings=table_settings_lines)
                if not tables:
                    tables = page.find_tables()

                if not tables:
                    logger.debug(f"Página {page_num}: sin tablas, se omite.")
                    continue

                sheet_name = f"Table {page_num}"
                ws = wb.create_sheet(title=sheet_name)
                sheets_created += 1

                # ── Recopilar TODAS las celdas de TODAS las tablas de la página ──
                all_cells = []
                for table in tables:
                    for row in table.rows:
                        for cell in row.cells:
                            if cell is not None:
                                all_cells.append(cell)  # (x0, y0, x1, y1)

                if not all_cells:
                    wb.remove(ws)
                    sheets_created -= 1
                    continue

                # ── Construir cuadrícula global para toda la página ──
                xs_raw = [c[0] for c in all_cells] + [c[2] for c in all_cells]
                ys_raw = [c[1] for c in all_cells] + [c[3] for c in all_cells]

                xs_grid = _cluster_coordinates(xs_raw, tolerance=X_TOLERANCE)
                ys_grid = _cluster_coordinates(ys_raw, tolerance=Y_TOLERANCE)

                cells_written: dict = {}
                merges_to_apply: set = set()

                for table in tables:
                    for row in table.rows:
                        for cell in row.cells:
                            if cell is None:
                                continue

                            x0, y0, x1, y1 = cell

                            col_start = _find_grid_index(x0, xs_grid, X_TOLERANCE)
                            col_end   = _find_grid_index(x1, xs_grid, X_TOLERANCE)
                            row_start = _find_grid_index(y0, ys_grid, Y_TOLERANCE)
                            row_end   = _find_grid_index(y1, ys_grid, Y_TOLERANCE)

                            # Convertir a coordenadas 1-based de Excel
                            excel_r1 = row_start + 1
                            excel_r2 = row_end      # inclusive end → row_end (no +1)
                            excel_c1 = col_start + 1
                            excel_c2 = col_end

                            # Sanity check
                            if excel_r2 < excel_r1 or excel_c2 < excel_c1:
                                continue

                            cell_id = (excel_r1, excel_c1)

                            # Extraer texto con padding para evitar artefactos de borde
                            crop_box = (
                                x0 + CROP_PADDING,
                                y0 + CROP_PADDING,
                                x1 - CROP_PADDING,
                                y1 - CROP_PADDING,
                            )
                            cropped = page.crop(crop_box)
                            raw_text = cropped.extract_text()
                            text_val = _clean_pdf_text(raw_text.strip()) if raw_text else ""

                            # Resolución de conflictos: si la celda ya fue escrita,
                            # conservar el texto con mayor cantidad de contenido.
                            # Esto protege contra duplicados exactos del PDF (capas superpuestas)
                            # sin perder contenido de páginas con múltiples tablas detectadas.
                            if cell_id in cells_written:
                                existing_text = cells_written[cell_id]
                                # Texto idéntico → duplicado real, descartar
                                if text_val.strip().lower() == existing_text.strip().lower():
                                    continue
                                # Si el nuevo tiene más contenido, reemplazar
                                if len(text_val.strip()) > len(existing_text.strip()):
                                    cells_written[cell_id] = text_val
                                    # (re-escribir la celda Excel con el texto ganador)
                                else:
                                    continue  # existente ya tiene más contenido
                            else:
                                cells_written[cell_id] = text_val

                            # Buscar color en page.rects usando coordenadas invertidas de Y
                            color_val = None
                            cc_x = (x0 + x1) / 2
                            cc_y = (y0 + y1) / 2
                            cc_y_inv = page.height - cc_y
                            
                            best_rect_area = float('inf')
                            for r in page.rects:
                                if not r.get("non_stroking_color"):
                                    continue
                                rx0 = r["x0"]
                                ry0 = r["y0"]
                                rx1 = r["x1"]
                                ry1 = r["y1"]
                                if (rx0 - 0.5 <= cc_x <= rx1 + 0.5) and (ry0 - 0.5 <= cc_y_inv <= ry1 + 0.5):
                                    area = (rx1 - rx0) * (ry1 - ry0)
                                    if area < best_rect_area:
                                        best_rect_area = area
                                        color_val = r["non_stroking_color"]

                            from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
                            thin = Side(border_style="thin", color="000000")
                            
                            c_obj = ws.cell(row=excel_r1, column=excel_c1, value=text_val)
                            c_obj.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            c_obj.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            
                            is_header = False
                            if text_val and text_val.strip().upper() in ["HORA", "LUNES", "MARTES", "MIÉRCOLES", "MIERCOLES", "JUEVES", "VIERNES", "SÁBADO", "SABADO"]:
                                c_obj.font = Font(bold=True)
                                is_header = True

                            hex_color = pdf_color_to_hex(color_val)
                            if hex_color:
                                if hex_color.upper() != "FFFFFF":
                                    c_obj.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                            elif is_header:
                                c_obj.fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

                            # Registrar fusión si la celda ocupa más de 1 fila o columna
                            if excel_r2 > excel_r1 or excel_c2 > excel_c1:
                                merges_to_apply.add((excel_r1, excel_c1, excel_r2, excel_c2))

                # Aplicar fusiones al final para evitar conflictos
                for r1, c1, r2, c2 in merges_to_apply:
                    try:
                        ws.merge_cells(
                            start_row=r1, start_column=c1,
                            end_row=r2,   end_column=c2
                        )
                    except Exception as merge_err:
                        logger.debug(f"No se pudo aplicar merge ({r1},{c1})-({r2},{c2}): {merge_err}")

    except Exception as e:
        logger.error(f"Error convirtiendo PDF '{pdf_path}': {e}", exc_info=True)
        raise RuntimeError(f"Error al convertir PDF: {e}") from e

    if sheets_created == 0:
        raise ValueError(
            f"El PDF '{Path(pdf_path).name}' no contiene tablas detectables. "
            "Verifica que el archivo sea una oferta académica válida."
        )

    # Guardar en archivo temporal
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", prefix="pdf_converted_")
    os.close(tmp_fd)

    try:
        wb.save(tmp_path)
    except Exception as save_err:
        os.unlink(tmp_path)
        raise RuntimeError(f"No se pudo guardar el XLSX temporal: {save_err}") from save_err

    logger.info(
        f"PDF convertido exitosamente: '{Path(pdf_path).name}' → "
        f"{sheets_created} hoja(s) → '{tmp_path}'"
    )
    return tmp_path


def is_pdf_file(filename: str, content_bytes: Optional[bytes] = None) -> bool:
    """
    Determina si un archivo es un PDF.

    Comprueba primero la firma mágica del contenido (si se proporciona),
    luego la extensión del nombre de archivo.
    """
    PDF_MAGIC = b"%PDF"
    if content_bytes and len(content_bytes) >= 4:
        return content_bytes[:4] == PDF_MAGIC
    return filename.lower().endswith(".pdf")
