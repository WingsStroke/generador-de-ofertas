"""
MergedCellHandler - Maneja celdas fusionadas en archivos Excel
"""
from typing import Dict, Optional, Any
from openpyxl.utils import get_column_letter


class MergedCellHandler:
    """
    Maneja celdas fusionadas que contienen datos distribuidos en múltiples filas/columnas.
    Útil para catálogos inline donde una celda fusionada contiene información que aplica
    a varias filas.
    """
    
    def __init__(self, sheet):
        self.sheet = sheet
        self.merged_ranges = list(sheet.merged_cells.ranges)
        self.merged_map = self._build_merged_map()
    
    def _build_merged_map(self) -> Dict:
        """
        Crea mapa de celdas fusionadas.
        
        Returns:
            Dict con clave (row, col) -> info de celda fusionada
        """
        merged_map = {}
        
        for merged_range in self.merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            
            # El valor real está en la celda maestra (superior-izquierda)
            master_value = self.sheet.cell(row=min_row, column=min_col).value
            
            # Mapear todas las celdas del rango
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    merged_map[(r, c)] = {
                        'value': master_value,
                        'is_master': (r == min_row and c == min_col),
                        'master_row': min_row,
                        'master_col': min_col,
                        'min_row': min_row,
                        'max_row': max_row,
                        'min_col': min_col,
                        'max_col': max_col,
                        'range': str(merged_range)
                    }
        
        return merged_map
    
    def get_effective_value(self, row: int, col: int) -> Any:
        """
        Retorna el valor efectivo de una celda, considerando celdas fusionadas.
        
        Args:
            row: Fila de la celda
            col: Columna de la celda
            
        Returns:
            Valor de la celda (propio o de la celda maestra si está fusionada)
        """
        if (row, col) in self.merged_map:
            return self.merged_map[(row, col)]['value']
        return self.sheet.cell(row=row, column=col).value
    
    def is_merged_cell(self, row: int, col: int) -> bool:
        """Verifica si una celda está dentro de un rango fusionado."""
        return (row, col) in self.merged_map
    
    def get_merged_info(self, row: int, col: int) -> Optional[Dict]:
        """
        Retorna información completa de la celda fusionada.
        
        Returns:
            Dict con info o None si no está fusionada
        """
        return self.merged_map.get((row, col))
    
    def find_merged_ranges_in_area(self, min_row: int, max_row: int, 
                                   min_col: int, max_col: int) -> list:
        """
        Encuentra todos los rangos fusionados que intersectan un área.
        
        Útil para detectar celdas fusionadas en el área del catálogo.
        """
        result = []
        for merged_range in self.merged_ranges:
            m_min_col, m_min_row, m_max_col, m_max_row = merged_range.bounds
            
            # Verificar intersección
            if (m_min_row <= max_row and m_max_row >= min_row and
                m_min_col <= max_col and m_max_col >= min_col):
                result.append({
                    'range': str(merged_range),
                    'min_row': m_min_row,
                    'max_row': m_max_row,
                    'min_col': m_min_col,
                    'max_col': m_max_col,
                    'value': self.sheet.cell(row=m_min_row, column=m_min_col).value
                })
        
        return result
    
    def __repr__(self):
        return f"MergedCellHandler({len(self.merged_ranges)} rangos fusionados)"
